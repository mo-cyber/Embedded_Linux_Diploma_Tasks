import openpyxl
path = 'E:/Embedded Linux Diploma/1- Python/04_advanced cont/Tasks/header_files.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook["employees"]
columns = ['name', 'job', 'salary']
number = 2
while(True):
    task = input("What do you want to do [add, print, remove]: ")
    if (task == 'add'):
        employee = [str(number),]
        for i in range(3):
            data = input(f"Enter {columns[i]}: ")
            employee.append(data)
        sheet.append(employee)
        number += 1
    elif (task == 'print'):
        id1 = input("Enter the ID you want to print: ")
        for row in sheet.iter_rows():
            if(row[0].value == id1):
                print(row[0].value, row[1].value, row[2].value, row[3].value)
                break
    elif (task == 'remove'):
        id2 = input("Enter the ID you want to remove: ")
        counter = 1
        for row in sheet.iter_rows():
            if(row[0].value == id2):
                sheet.delete_rows(counter)
                break
            counter += 1
    workbook.save(path)
    workbook.close()
        

